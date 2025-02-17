# pyinstaller --onefile --noconsole docu24.py


import os, sys, time
import ctypes
# sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "selenium")))

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import load_workbook

def show_alert(message):
    ctypes.windll.user32.MessageBoxW(0, message, "오류", 0x10)
def show_result(message):
    ctypes.windll.user32.MessageBoxW(0, message, "완료", 0x40)

# 실행 파일이 있는 경로 찾기
if getattr(sys, 'frozen', False):
    # PyInstaller로 빌드된 경우 (실행 파일 기준)
    base_path = os.path.dirname(sys.executable)
else:
    # 개발 중 (스크립트 직접 실행)
    base_path = os.path.dirname(__file__)

file_path = os.path.join(base_path, "docu24.xlsx")
workbook = load_workbook(file_path)
ws = workbook['발송리스트']

user_type = ws['H2'].value
user_id = ws['H3'].value
password = ws['H4'].value

chrome_options = Options()
chrome_options.add_experimental_option("detach", True)  # 브라우저 유지 옵션
# chrome_options.add_argument("--window-size=1920,1080")  # 예: 1920x1080 해상도
chrome_options.add_argument("--window-position=0,0")

# ChromeDriver 경로 설정 (크롬드라이버 설치 경로)
driver_path = os.path.join(base_path, "chromedriver.exe")
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

try:
    # 웹사이트 열고 로그인하기
    driver.get("https://docu.gdoc.go.kr/cmm/main/loginForm.do")
    wait = WebDriverWait(driver, 10)
    
    if user_type == '법인/단체사용자':        
        enterprise = wait.until(EC.presence_of_element_located((By.ID, "entrprsHref")))
        enterprise.click()
        login_radio = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "label[for='st_rdo3']")))
        login_radio.click()
    elif user_type == '개인사용자':
        personal = wait.until(EC.presence_of_element_located((By.ID, "gnrlHref")))
        personal.click()
    else:
        raise Exception("잘못된 user_type 입니다.")
    
    if not user_id or not password:
        raise Exception("계정 정보를 엑셀에 입력해주세요.")

    login_id = wait.until(EC.presence_of_element_located((By.ID, "id")))
    login_id.send_keys(user_id)
    login_password = wait.until(EC.presence_of_element_located((By.ID, "password")))
    login_password.send_keys(password)
    login_btn = wait.until(EC.presence_of_element_located((By.ID, "loginBtn")))
    login_btn.click()
    
    main_window = driver.current_window_handle
    popup_wait = WebDriverWait(driver, 2)
    try:
        popup_wait.until(lambda driver: len(driver.window_handles) > 1)

        # 팝업이 있으면 팝업 닫기
        for handle in driver.window_handles:
            driver.switch_to.window(handle)
            if handle != main_window:
                driver.close()  # 팝업 닫기
    except TimeoutException:
        print("팝업 없음")
        
    driver.switch_to.window(main_window)
    
    # 문서 자동발송하기
    total_rows = ws.max_row
    complete_cnt = 0
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=total_rows)):
        try:
            docu_title = row[0].value
            target = row[1].value
            docu_num1 = row[2].value
            docu_num2 = row[3].value
            if not docu_title or not target:
                row[4].value = "필수값 누락"
                continue

            #보낸문서함에서 발송할 문서 찾아서 재작성
            driver.get("https://docu.gdoc.go.kr/doc/snd/sendDocList.do")
            search_input = wait.until(EC.presence_of_element_located((By.ID, "defaultSearchWord")))
            search_input.send_keys(docu_title)
            search_btn = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".open_searcher button")))
            search_btn.click()
    
            loading_element = wait.until(EC.invisibility_of_element_located((By.ID, "loading")))
            time.sleep(0.5)

            sent_historys = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "tbody tr:nth-child(1) td")))
            #이전에 발송 문서 없을 경우
            if len(sent_historys) < 2:
                row[4].value = "대상 문서 없음"
                continue
            
            latest_sent = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "tbody tr:nth-child(1) td:nth-child(2) a")))
            latest_sent.click()
            rewrite_btn = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".table_btn_wrap ul li:nth-child(2) button")))
            rewrite_btn.click()
            time.sleep(0.3)
            load_file_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#jconfirm-buttons-bottom button:nth-child(2)")))
            load_file_btn.click()
            
            #문서 작성
            loading_element = wait.until(EC.invisibility_of_element_located((By.ID, "loading")))
            time.sleep(0.5)

            check_items = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".chk_wrap div label")))
            for item in check_items:
                driver.execute_script("arguments[0].click();", item)
                
            popup_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#jconfirm-buttons-bottom button:nth-child(1)")))
            popup_btn.click()
            
            #받는기관 선택
            recipient_search = wait.until(EC.presence_of_element_located((By.ID, "ldapSearch")))
            recipient_search.click()
            search_recipient_name_input = wait.until(EC.element_to_be_clickable((By.ID, "searchOrgNm")))
            search_recipient_name_input.send_keys(target)
            search_btn = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#Form button")))
            search_btn.click()
            
            loading_element = wait.until(EC.invisibility_of_element_located((By.ID, "loading")))
            time.sleep(1)

            search_results = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#ldaps tbody tr:nth-child(1) td")))
            if len(search_results) < 2:
                row[4].value = "대상 기관 없음"
                continue
            
            search_results_trs = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#ldaps tbody tr")))
            target_corp = ''
            for idx, tr in enumerate(search_results_trs):
                corp_name = tr.find_elements(By.TAG_NAME, "td")[1].text
                print(corp_name)
                if corp_name == target:
                    target_corp = corp_name
                    select_btn = tr.find_element(By.CSS_SELECTOR, "td:nth-child(1) button")
                    select_btn.click()
                    break
            
            loading_element = wait.until(EC.invisibility_of_element_located((By.ID, "loading")))
            time.sleep(0.5)

            # 커스텀 문서번호 있을 경우
            if docu_num1:
                #문서번호 변경
                regnum_checkbox = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "label[for='regnumAutoChk']")))
                regnum_checkbox.click()
                time.sleep(0.5)

                popup_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#jconfirm-buttons-bottom button:nth-child(2)")))
                popup_btn.click()
                
                loading_element = wait.until(EC.invisibility_of_element_located((By.ID, "loading")))
                time.sleep(0.5)

                regnum1_input = wait.until(EC.element_to_be_clickable((By.ID, "regnum1")))
                regnum1_input.clear()
                regnum1_input.send_keys(docu_num1)
                regnum2_input = wait.until(EC.element_to_be_clickable((By.ID, "regnum2")))
                regnum2_input.clear()
                regnum2_input.send_keys(docu_num2)
            
            send_btn =  wait.until(EC.element_to_be_clickable((By.ID, "sendDoc")))
            send_btn.click()
            time.sleep(0.5)
        
            send_confirm_btn = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#jconfirm-buttons-top button:nth-child(2)")))
            loading_element = wait.until(EC.invisibility_of_element_located((By.ID, "loading")))

            time.sleep(0.5)
            if send_confirm_btn.is_displayed():
                time.sleep(0.3)
                print("Button is visible and can be interacted with.")
                send_confirm_btn.click()
            else:
                send_confirm_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#jconfirm-buttons-bottom button:nth-child(2)")))
                time.sleep(0.5)
                send_confirm_btn.click()
            time.sleep(0.5)

            popup_bottom_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".jconfirm-doc24_alert #jconfirm-buttons-bottom button:nth-child(2)")))
            popup_bottom_btn.click()
            
            time.sleep(3)
            row[4].value = "발송 완료"
            row[5].value = target_corp
            complete_cnt += 1
            if (idx + 1) % 10 == 0:
                workbook.save(file_path)
        except:
            row[4].value = "오류 발생"
    workbook.save(file_path)
    show_result(f"{complete_cnt}건 발송 완료. 발송결과는 엑셀 문서를 확인해주세요.")
except Exception as e:
    show_alert(f"에러 발생: {e}")
finally:
    print(driver)
    # # 브라우저 닫기
    driver.quit()
